import win32com.client as win32
import pandas as pd
import openpyxl
import sys


def read_email_list():
    outlook = win32.Dispatch("outlook.application")
    wb = openpyxl.load_workbook("c:\\users\\u6zn\\desktop\\write_off_mail_list.xlsx")
    sheet = wb["Sheet1"]
    recipients_list = {}

    for r in range(2, 88):
        location = sheet.cell(row=r, column=1).value
        name = sheet.cell(row=r, column=2).value
        email = sheet.cell(row=r, column=3).value
        supv = sheet.cell(row=r, column=4).value
        atch1 = sheet.cell(row=r, column=5).value
        atch2 = sheet.cell(row=r, column=6).value
        subj = sheet.cell(row=r, column=7).value
        recipients_list[location] = [name, email, supv, atch1, atch2, subj]
    for location in recipients_list.items():
        print(
            f"Store Location: {location[0]} Subject: {location[1][5]} Store Keeper's First Name(s): {location[1][0]} SK email: {location[1][1]} Supervisor: {location[1][2]} Attachment1: {location[1][3]} Attachment2: {location[1][4]}"
        )
        mail = outlook.CreateItem(0)
        mail.To = location[1][1]
        mail.CC = location[1][2]
        mail.Subject = location[1][5]
        mail.Attachments.Add(location[1][3])
        mail.Attachments.Add(location[1][4])
        mail.HtmlBody = (
            "<html xmlns:v=\"urn:schemas-microsoft-com:vml\" xmlns:o=\"urn:schemas-microsoft-com:office:office\" xmlns:w=\"urn:schemas-microsoft-com:office:word\" xmlns:x=\"urn:schemas-microsoft-com:office:excel\" xmlns:m=\"http://schemas.microsoft.com/office/2004/12/omml\" xmlns=\"http://www.w3.org/TR/REC-html40\"><head><meta http-equiv=Content-Type content=\"text/html; charset=us-ascii\"><meta name=Generator content=\"Microsoft Word 15 (filtered medium)\"><style><!--\n"
            "/* Font Definitions */\n"
            "@font-face\n"
            "	{font-family:\"Cambria Math\";\n"
            "	panose-1:2 4 5 3 5 4 6 3 2 4;}\n"
            "@font-face\n"
            "	{font-family:Calibri;\n"
            "	panose-1:2 15 5 2 2 2 4 3 2 4;}\n"
            "@font-face\n"
            "	{font-family:Verdana;\n"
            "	panose-1:2 11 6 4 3 5 4 4 2 4;}\n"
            "/* Style Definitions */\n"
            "p.MsoNormal, li.MsoNormal, div.MsoNormal\n"
            "	{margin:0in;\n"
            "	margin-bottom:.0001pt;\n"
            "	font-size:11.0pt;\n"
            "	font-family:\"Calibri\",sans-serif;}\n"
            "a:link, span.MsoHyperlink\n"
            "	{mso-style-priority:99;\n"
            "	color:#0563C1;\n"
            "	text-decoration:underline;}\n"
            "a:visited, span.MsoHyperlinkFollowed\n"
            "	{mso-style-priority:99;\n"
            "	color:#954F72;\n"
            "	text-decoration:underline;}\n"
            "p.MsoListParagraph, li.MsoListParagraph, div.MsoListParagraph\n"
            "	{mso-style-priority:34;\n"
            "	margin-top:0in;\n"
            "	margin-right:0in;\n"
            "	margin-bottom:0in;\n"
            "	margin-left:.5in;\n"
            "	margin-bottom:.0001pt;\n"
            "	font-size:11.0pt;\n"
            "	font-family:\"Calibri\",sans-serif;}\n"
            "span.EmailStyle18\n"
            "	{mso-style-type:personal;\n"
            "	font-family:\"Calibri\",sans-serif;\n"
            "	color:windowtext;}\n"
            "span.EmailStyle19\n"
            "	{mso-style-type:personal-reply;\n"
            "	font-family:\"Calibri\",sans-serif;\n"
            "	color:#1F497D;}\n"
            ".MsoChpDefault\n"
            "	{mso-style-type:export-only;\n"
            "	font-size:10.0pt;}\n"
            "@page WordSection1\n"
            "	{size:8.5in 11.0in;\n"
            "	margin:1.0in 1.0in 1.0in 1.0in;}\n"
            "div.WordSection1\n"
            "	{page:WordSection1;}\n"
            "/* List Definitions */\n"
            "@list l0\n"
            "	{mso-list-id:642931991;\n"
            "	mso-list-type:hybrid;\n"
            "	mso-list-template-ids:-1664984492 67698703 67698713 67698715 67698703 67698713 67698715 67698703 67698713 67698715;}\n"
            "@list l0:level1\n"
            "	{mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l0:level2\n"
            "	{mso-level-number-format:alpha-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l0:level3\n"
            "	{mso-level-number-format:roman-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:right;\n"
            "	text-indent:-9.0pt;}\n"
            "@list l0:level4\n"
            "	{mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l0:level5\n"
            "	{mso-level-number-format:alpha-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l0:level6\n"
            "	{mso-level-number-format:roman-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:right;\n"
            "	text-indent:-9.0pt;}\n"
            "@list l0:level7\n"
            "	{mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l0:level8\n"
            "	{mso-level-number-format:alpha-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l0:level9\n"
            "	{mso-level-number-format:roman-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:right;\n"
            "	text-indent:-9.0pt;}\n"
            "@list l1\n"
            "	{mso-list-id:657154297;\n"
            "	mso-list-type:hybrid;\n"
            "	mso-list-template-ids:-1265058316 67698703 67698713 67698715 67698703 67698713 67698715 67698703 67698713 67698715;}\n"
            "@list l1:level1\n"
            "	{mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l1:level2\n"
            "	{mso-level-number-format:alpha-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l1:level3\n"
            "	{mso-level-number-format:roman-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:right;\n"
            "	text-indent:-9.0pt;}\n"
            "@list l1:level4\n"
            "	{mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l1:level5\n"
            "	{mso-level-number-format:alpha-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l1:level6\n"
            "	{mso-level-number-format:roman-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:right;\n"
            "	text-indent:-9.0pt;}\n"
            "@list l1:level7\n"
            "	{mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l1:level8\n"
            "	{mso-level-number-format:alpha-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:left;\n"
            "	text-indent:-.25in;}\n"
            "@list l1:level9\n"
            "	{mso-level-number-format:roman-lower;\n"
            "	mso-level-tab-stop:none;\n"
            "	mso-level-number-position:right;\n"
            "	text-indent:-9.0pt;}\n"
            "ol\n"
            "	{margin-bottom:0in;}\n"
            "ul\n"
            "	{margin-bottom:0in;}\n"
            "--></style><!--[if gte mso 9]><xml>\n"
            "<o:shapedefaults v:ext=\"edit\" spidmax=\"1026\" />\n"
            "</xml><![endif]--><!--[if gte mso 9]><xml>\n"
            "<o:shapelayout v:ext=\"edit\">\n"
            "<o:idmap v:ext=\"edit\" data=\"1\" />\n"
            f"</o:shapelayout></xml><![endif]--></head><body lang=EN-US link=\"#0563C1\" vlink=\"#954F72\"><div class=WordSection1><p class=MsoNormal>{location[1][0]}-<o:p></o:p></p><p class=MsoNormal><span style='color:#1F497D'><o:p>&nbsp;</o:p></span></p><p class=MsoNormal>Please review the attached spreadsheet for your location&#8217;s Obsolete, Inactive, &amp; Excess Inventory Review.<o:p></o:p></p><p class=MsoNormal>Select one of the following options in the <i><u>Decision</u></i> column for each listed item&#8217;s <span style='background:yellow;mso-highlight:yellow'>Review Quantity</span>:<o:p></o:p></p><p class=MsoNormal><o:p>&nbsp;</o:p></p><p class=MsoListParagraph style='text-indent:-.25in;mso-list:l1 level1 lfo2'><![if !supportLists]><span style='mso-list:Ignore'>1.<span style='font:7.0pt \"Times New Roman\"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; </span></span><![endif]>Retain All<o:p></o:p></p><p class=MsoListParagraph style='text-indent:-.25in;mso-list:l1 level1 lfo2'><![if !supportLists]><span style='mso-list:Ignore'>2.<span style='font:7.0pt \"Times New Roman\"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; </span></span><![endif]>Retain None<o:p></o:p></p><p class=MsoListParagraph style='text-indent:-.25in;mso-list:l1 level1 lfo2'><![if !supportLists]><span style='mso-list:Ignore'>3.<span style='font:7.0pt \"Times New Roman\"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; </span></span><![endif]>Transfer or Write-Off Review Quantity<o:p></o:p></p><p class=MsoListParagraph><o:p>&nbsp;</o:p></p><p class=MsoNormal>For each item that will have a quantity retained, please select one of the following options in the <i><u>Reason for Retention</u></i> column:<o:p></o:p></p><p class=MsoNormal><o:p>&nbsp;</o:p></p><p class=MsoListParagraph style='text-indent:-.25in;mso-list:l0 level1 lfo4'><![if !supportLists]><span style='mso-list:Ignore'>1.<span style='font:7.0pt \"Times New Roman\"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; </span></span><![endif]>Emergency Spare<o:p></o:p></p><p class=MsoListParagraph style='text-indent:-.25in;mso-list:l0 level1 lfo4'><![if !supportLists]><span style='mso-list:Ignore'>2.<span style='font:7.0pt \"Times New Roman\"'>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; </span></span><![endif]>Future Project Use<o:p></o:p></p><p class=MsoNormal><o:p>&nbsp;</o:p></p><p class=MsoNormal>**<b>Please refer to the attached PDF for assistance on populating the </b><i><u>Decision</u></i><b> &amp; </b><i><u>Reason for Retention</u></i><b> columns</b>**<o:p></o:p></p><p class=MsoNormal><o:p>&nbsp;</o:p></p><p class=MsoNormal>For all items listed on your spreadsheet, please fill in the column <i><u>Decision Maker&#8217;s Name</u></i>.<o:p></o:p></p><p class=MsoNormal><o:p>&nbsp;</o:p></p><p class=MsoNormal>Once you have populated the <i>Decision Maker&#8217;s Name,</i> <i>Decision</i>, and <i>Reason for Retention </i>(if retaining material) for all items listed, please save your spreadsheet and email it back to me. The deadline for returning your completed spreadsheet is <b><u><span style='font-size:14.0pt'>COB Wednesday, July 22<sup>th</sup></span></u></b>. <o:p></o:p></p><p class=MsoNormal><o:p>&nbsp;</o:p></p><p class=MsoNormal>Please do not execute any actual transactions for this process until instructed to do so. Transactions cannot be completed until approved by Senior Management. If you need any help or have any questions about the above process, please don&#8217;t hesitate to email or call me. My contact information can be found at the end of this email.<o:p></o:p></p><p class=MsoNormal><o:p>&nbsp;</o:p></p><p class=MsoNormal><o:p>&nbsp;</o:p></p><p class=MsoNormal><span style='font-size:10.0pt;font-family:\"Verdana\",sans-serif;color:#1F497D'>Thank you,<o:p></o:p></span></p><p class=MsoNormal><span style='font-size:10.0pt;font-family:\"Verdana\",sans-serif;color:#1F497D'><o:p>&nbsp;</o:p></span></p><p class=MsoNormal><b><span style='font-family:\"Arial\",sans-serif;color:#7F7F7F'>Kevin Vaughan<o:p></o:p></span></b></p><p class=MsoNormal style='text-autospace:none'><span style='font-family:\"Arial\",sans-serif;color:gray'>Material Management Analyst<o:p></o:p></span></p><p class=MsoNormal style='text-autospace:none'><b><span style='font-family:\"Arial\",sans-serif;color:#717073'>Oncor | Supply Chain Management<o:p></o:p></span></b></p><p class=MsoNormal style='text-autospace:none'><span style='font-family:\"Arial\",sans-serif;color:#717073;mso-fareast-language:ZH-CN'>15101 Trinity Blvd. <o:p></o:p></span></p><p class=MsoNormal style='text-autospace:none'><span style='font-family:\"Arial\",sans-serif;color:#717073;mso-fareast-language:ZH-CN'>Fort Worth, TX 76155<o:p></o:p></span></p><p class=MsoNormal style='text-autospace:none'><span style='font-family:\"Arial\",sans-serif;color:gray'><o:p>&nbsp;</o:p></span></p><p class=MsoNormal><span style='font-family:\"Arial\",sans-serif;color:gray'>Tel: 817.359.2310<o:p></o:p></span></p><p class=MsoNormal><span style='font-family:\"Arial\",sans-serif;color:#7F7F7F'>Cell: 202.906.9503<o:p></o:p></span></p><p class=MsoNormal><a href=\"mailto:Kevin.Vaughan2@oncor.com\">Kevin.Vaughan2@oncor.com</a><span style='color:#2E74B5'><o:p></o:p></span></p><p class=MsoNormal><span style='color:#1F497D'><o:p>&nbsp;</o:p></span></p><p class=MsoNormal><b><span style='font-family:\"Arial\",sans-serif;color:#2E74B5'>oncor.com<o:p></o:p></span></b></p><p class=MsoNormal><o:p>&nbsp;</o:p></p></div></body></html>")

        # mail.Send() ## uncomment or it won't work


read_email_list()

# def Emailer(text, subject, recipient, attachment):

#     outlook = win32.Dispatch('outlook.application')
#     mail = outlook.CreateItem(0)
#     mail.To = recipient
#     mail.Subject = subject
#     mail.HtmlBody = text
#     mail.Attachments.Add(attachment)
#     mail.Display(True)


# working_list_test = pd.read_excel("c:\\users\\u6zn\\desktop\\test_list.xlsx")

# for row in working_list_test:
#     print(working_list_test['LOCATION'], working_list_test['STOREKEEPER'], working_list_test['SUPERVISOR'], working_list_test['FIRSTNAME'])
# attachment = "c:\\users\\u6zn\\desktop\\test_list.xlsx"

# Emailer("test body", "test subject", "test.rec.poc", attachment)
# for row in working_list_test:
#     Emailer("test body", "test subject", email)
