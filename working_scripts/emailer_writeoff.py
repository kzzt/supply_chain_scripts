import win32com.client as win32
import pandas as pd
import openpyxl
import sys


def read_email_list():
    outlook = win32.Dispatch("outlook.application")
    wb = openpyxl.load_workbook(
        "c:\\users\\u6zn\\desktop\\write_off_mail_list_JANUARY_REMINDER.xlsx"
    )
    sheet = wb["Sheet1"]
    recipients_list = {}

    usage_number = "0101010101"
    wo_number = "3214321"

    for r in range(2, 88):
        location = sheet.cell(row=r, column=1).value
        name = sheet.cell(row=r, column=2).value
        email = sheet.cell(row=r, column=3).value
        supv = sheet.cell(row=r, column=4).value
        atch1 = sheet.cell(row=r, column=5).value
        atch2 = sheet.cell(row=r, column=6).value
        subj = sheet.cell(row=r, column=7).value
        recipients_list[location] = [name, email, supv, atch1, atch2, subj]
    for location in recipients_list.items():
        print(
            f"Store Location: {location[0]} Subject: {location[1][5]} Store Keeper's First Name(s): {location[1][0]} SK email: {location[1][1]} Supervisor: {location[1][2]} Attachment1: {location[1][3]} Attachment2: {location[1][4]}"
        )
        mail = outlook.CreateItem(0)
        mail.To = location[1][1]
        # mail.CC = location[1][2]
        mail.Subject = location[1][5]
        # mail.Attachments.Add(location[1][3])
        # mail.Attachments.Add(location[1][4])
        mail.HtmlBody = (
            '<html xmlns:v="urn:schemas-microsoft-com:vml"\n'
            'xmlns:o="urn:schemas-microsoft-com:office:office"\n'
            'xmlns:w="urn:schemas-microsoft-com:office:word"\n'
            'xmlns:m="http://schemas.microsoft.com/office/2004/12/omml"\n'
            'xmlns="http://www.w3.org/TR/REC-html40">\n'
            "\n"
            "<head>\n"
            '<meta http-equiv=Content-Type content="text/html; charset=windows-1252">\n'
            "<meta name=ProgId content=Word.Document>\n"
            '<meta name=Generator content="Microsoft Word 15">\n'
            '<meta name=Originator content="Microsoft Word 15">\n'
            '<link rel=File-List href="email%20template%20test_files/filelist.xml">\n'
            '<link rel=Edit-Time-Data href="email%20template%20test_files/editdata.mso">\n'
            '<link rel=themeData href="email%20template%20test_files/themedata.thmx">\n'
            "<link rel=colorSchemeMapping\n"
            'href="email%20template%20test_files/colorschememapping.xml">\n'
            "<style>\n"
            "<!--\n"
            " /* Font Definitions */\n"
            " @font-face\n"
            '	{font-family:"Cambria Math";\n'
            "	panose-1:2 4 5 3 5 4 6 3 2 4;\n"
            "	mso-font-charset:1;\n"
            "	mso-generic-font-family:roman;\n"
            "	mso-font-format:other;\n"
            "	mso-font-pitch:variable;\n"
            "	mso-font-signature:0 0 0 0 0 0;}\n"
            "@font-face\n"
            "	{font-family:Calibri;\n"
            "	panose-1:2 15 5 2 2 2 4 3 2 4;\n"
            "	mso-font-charset:0;\n"
            "	mso-generic-font-family:swiss;\n"
            "	mso-font-pitch:variable;\n"
            "	mso-font-signature:-536858881 -1073732485 9 0 511 0;}\n"
            "@font-face\n"
            "	{font-family:Verdana;\n"
            "	panose-1:2 11 6 4 3 5 4 4 2 4;\n"
            "	mso-font-charset:0;\n"
            "	mso-generic-font-family:swiss;\n"
            "	mso-font-pitch:variable;\n"
            "	mso-font-signature:-1610610945 1073750107 16 0 415 0;}\n"
            " /* Style Definitions */\n"
            " p.MsoNormal, li.MsoNormal, div.MsoNormal\n"
            "	{mso-style-unhide:no;\n"
            "	mso-style-qformat:yes;\n"
            '	mso-style-parent:"";\n'
            "	margin:0in;\n"
            "	margin-bottom:.0001pt;\n"
            "	mso-pagination:widow-orphan;\n"
            "	font-size:11.0pt;\n"
            '	font-family:"Calibri",sans-serif;\n'
            "	mso-fareast-font-family:Calibri;\n"
            "	mso-fareast-theme-font:minor-latin;}\n"
            "a:link, span.MsoHyperlink\n"
            "	{mso-style-noshow:yes;\n"
            "	mso-style-priority:99;\n"
            "	color:#0563C1;\n"
            "	text-decoration:underline;\n"
            "	text-underline:single;}\n"
            "a:visited, span.MsoHyperlinkFollowed\n"
            "	{mso-style-noshow:yes;\n"
            "	mso-style-priority:99;\n"
            "	color:#954F72;\n"
            "	text-decoration:underline;\n"
            "	text-underline:single;}\n"
            "span.EmailStyle17\n"
            "	{mso-style-type:personal;\n"
            "	mso-style-noshow:yes;\n"
            "	mso-style-unhide:no;\n"
            '	font-family:"Calibri",sans-serif;\n'
            "	mso-ascii-font-family:Calibri;\n"
            "	mso-hansi-font-family:Calibri;\n"
            "	mso-bidi-font-family:Calibri;\n"
            "	color:windowtext;}\n"
            ".MsoChpDefault\n"
            "	{mso-style-type:export-only;\n"
            "	mso-default-props:yes;\n"
            "	font-size:10.0pt;\n"
            "	mso-ansi-font-size:10.0pt;\n"
            "	mso-bidi-font-size:10.0pt;}\n"
            "@page WordSection1\n"
            "	{size:8.5in 11.0in;\n"
            "	margin:1.0in 1.0in 1.0in 1.0in;\n"
            "	mso-header-margin:.5in;\n"
            "	mso-footer-margin:.5in;\n"
            "	mso-paper-source:0;}\n"
            "div.WordSection1\n"
            "	{page:WordSection1;}\n"
            "-->\n"
            "</style>\n"
            "<!--[if gte mso 10]>\n"
            "<style>\n"
            " /* Style Definitions */\n"
            " table.MsoNormalTable\n"
            '	{mso-style-name:"Table Normal";\n'
            "	mso-tstyle-rowband-size:0;\n"
            "	mso-tstyle-colband-size:0;\n"
            "	mso-style-noshow:yes;\n"
            "	mso-style-priority:99;\n"
            '	mso-style-parent:"";\n'
            "	mso-padding-alt:0in 5.4pt 0in 5.4pt;\n"
            "	mso-para-margin:0in;\n"
            "	mso-para-margin-bottom:.0001pt;\n"
            "	mso-pagination:widow-orphan;\n"
            "	font-size:10.0pt;\n"
            '	font-family:"Times New Roman",serif;}\n'
            "</style>\n"
            "<![endif]--><!--[if gte mso 9]><xml>\n"
            ' <o:shapedefaults v:ext="edit" spidmax="1026"/>\n'
            "</xml><![endif]--><!--[if gte mso 9]><xml>\n"
            ' <o:shapelayout v:ext="edit">\n'
            '  <o:idmap v:ext="edit" data="1"/>\n'
            " </o:shapelayout></xml><![endif]-->\n"
            "</head>\n"
            "\n"
            '<body lang=EN-US link="#0563C1" vlink="#954F72" style=\'tab-interval:.5in\'>\n'
            "\n"
            "\n"
            f"<p class=MsoNormal><span style='color:#1F4E79'>{location[1][0]},<o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F4E79'><o:p>&nbsp;</o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F4E79'>This is a reminder:<o:p>&nbsp;</o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F4E79'>Please proceed with issuing\n"
            f"Inventory Usage #{location[1][4]} from Work Order #{location[1][3]} for the 2020 Write-Off. Please\n"
            "pull the exact items and quantities on this Work Order from their respective\n"
            "stock locations at your storeroom and ship them to Gilbert Return Center.<o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F4E79'>When packaging your Write-Off\n"
            "material, please clearly mark it with “2020 Write-Off” so it can be easily\n"
            "identified by receiving personnel. If necessary, Mr. Luke Harvey is available to\n"
            "assist in shipping coordination. <o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F4E79'><o:p>&nbsp;</o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F4E79'>Please complete the\n"
            "transaction(s) by <b><u>COB January 29th, 2021</u></b> and confirm the\n"
            "completion of this task with an email directly to me and to GPCRETURNS@oncor.com  <o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F4E79'><o:p>&nbsp;</o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F4E79'>Please disregard if\n"
            "you have already completed the write off for your location.<o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F4E79'><o:p>&nbsp;</o:p></span></p>\n"
            "\n"
            '<p class=MsoNormal><span style=\'font-size:10.0pt;font-family:"Verdana",sans-serif;\n'
            "color:#1F4E79'>Thank you,</span></p>\n"
            "\n"
            "<p class=MsoNormal><o:p>&nbsp;</o:p></p>\n"
            "\n"
            "<p class=MsoNormal><b><span style='font-family:\"Verdana\",sans-serif;color:#7F7F7F'>Kevin\n"
            "Vaughan<o:p></o:p></span></b></p>\n"
            "\n"
            "<p class=MsoNormal style='text-autospace:none'><span style='font-family:\"Verdana\",sans-serif;\n"
            "color:gray'>Material Management Analyst<o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal style='text-autospace:none'><b><span style='font-family:\n"
            '"Verdana",sans-serif;color:#717073\'>Oncor | Supply Chain Management<o:p></o:p></span></b></p>\n'
            "\n"
            "<p class=MsoNormal style='text-autospace:none'><span style='font-family:\"Verdana\",sans-serif;\n"
            "color:#717073;mso-fareast-language:ZH-CN'>15101 Trinity Blvd. <o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal style='text-autospace:none'><span style='font-family:\"Verdana\",sans-serif;\n"
            "color:#717073;mso-fareast-language:ZH-CN'>Fort Worth, TX 76155<o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal style='text-autospace:none'><span style='font-family:\"Verdana\",sans-serif;\n"
            "color:gray'><o:p>&nbsp;</o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='font-family:\"Verdana\",sans-serif;color:gray'>Tel:\n"
            "817.359.2310<o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='font-family:\"Verdana\",sans-serif;color:#7F7F7F'>Cell:\n"
            "214.542.6626<o:p></o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='font-family:\"Verdana\",sans-serif'><a\n"
            'href="mailto:Kevin.Vaughan@oncor.com">Kevin.Vaughan@oncor.com</a><span\n'
            "style='color:#2E74B5'><o:p></o:p></span></span></p>\n"
            "\n"
            "<p class=MsoNormal><span style='color:#1F497D'><o:p>&nbsp;</o:p></span></p>\n"
            "\n"
            "<p class=MsoNormal><b><span style='font-family:\"Arial\",sans-serif;color:#2E74B5'>oncor.com<o:p></o:p></span></b></p>\n"
            "\n"
            "<p class=MsoNormal><o:p>&nbsp;</o:p></p>\n"
            "\n"
            "</div>\n"
            "\n"
            "</body>\n"
            "\n"
            "</html>\n"
        )

        mail.Send() ## uncomment or it won't work
        # mail.Display(False)


read_email_list()
